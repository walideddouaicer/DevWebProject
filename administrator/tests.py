from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from openpyxl import Workbook

from .models import AdminProfile
from student.models import StudentProfile
from teacher.models import Module, ModuleEnrollment


def build_roster_file(rows, headers=None):
    """Build an in-memory .xlsx roster for the import tests"""
    if headers is None:
        headers = ['Prénom', 'Nom', 'Email', 'ID Étudiant', 'Année', 'Département']
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    buffer.name = 'roster.xlsx'
    return buffer


class ImportStudentsTests(TestCase):
    def setUp(self):
        admin_user = User.objects.create_user(username='admin', password='adminpass')
        AdminProfile.objects.create(user=admin_user, admin_id='A001')
        self.client.login(username='admin', password='adminpass')

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')

    def test_import_creates_students_and_enrollments(self):
        roster = build_roster_file([
            ['Amina', 'Benali', 'amina.benali@example.com', 'E12345', 3, 'Informatique'],
            ['Youssef', 'El Amrani', 'youssef.elamrani@example.com', 'E12346', 4, 'Génie Civil'],
        ])

        response = self.client.post(reverse('administrator:import_students'), {
            'excel_file': roster,
            'module_id': self.module.id,
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['created']), 2)
        self.assertEqual(response.context['errors'], [])

        student = StudentProfile.objects.get(student_id='E12345')
        self.assertEqual(student.user.email, 'amina.benali@example.com')
        self.assertEqual(student.year_of_study, 3)
        self.assertTrue(
            ModuleEnrollment.objects.filter(student=student, module=self.module, is_active=True).exists()
        )

        # Generated password must log the student in
        password = response.context['created'][0]['password']
        username = response.context['created'][0]['username']
        self.client.logout()
        self.assertTrue(self.client.login(username=username, password=password))

    def test_import_skips_existing_email_and_student_id(self):
        existing_user = User.objects.create_user(
            username='amina', email='amina.benali@example.com', password='x'
        )
        StudentProfile.objects.create(
            user=existing_user, student_id='E99999', year_of_study=3, department='Informatique'
        )

        roster = build_roster_file([
            ['Amina', 'Benali', 'amina.benali@example.com', 'E12345', 3, 'Informatique'],  # duplicate email
            ['Sara', 'Idrissi', 'sara.idrissi@example.com', 'E99999', 5, 'Informatique'],  # duplicate student id
            ['Omar', 'Tazi', 'omar.tazi@example.com', 'E55555', 4, 'Informatique'],        # valid
        ])

        response = self.client.post(reverse('administrator:import_students'), {'excel_file': roster})

        self.assertEqual(len(response.context['created']), 1)
        self.assertEqual(len(response.context['skipped']), 2)
        self.assertTrue(StudentProfile.objects.filter(student_id='E55555').exists())

    def test_import_reports_invalid_rows(self):
        roster = build_roster_file([
            ['Lina', 'Berrada', 'lina.berrada@example.com', 'E11111', 7, 'Informatique'],  # bad year
            ['', 'Berrada', 'no.first.name@example.com', 'E22222', 3, 'Informatique'],      # missing value
        ])

        response = self.client.post(reverse('administrator:import_students'), {'excel_file': roster})

        self.assertEqual(response.context['created'], [])
        self.assertEqual(len(response.context['errors']), 2)
        self.assertFalse(User.objects.filter(email='lina.berrada@example.com').exists())

    def test_import_rejects_file_with_missing_columns(self):
        roster = build_roster_file(
            [['Amina', 'Benali']],
            headers=['Prénom', 'Nom'],
        )

        response = self.client.post(reverse('administrator:import_students'), {'excel_file': roster})

        self.assertFalse(response.context['import_done'])
        self.assertEqual(StudentProfile.objects.count(), 0)

    def test_template_download(self):
        response = self.client.get(reverse('administrator:student_import_template'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])


class ImportTeachersTests(TestCase):
    def setUp(self):
        admin_user = User.objects.create_user(username='admin', password='adminpass')
        AdminProfile.objects.create(user=admin_user, admin_id='A001')
        self.client.login(username='admin', password='adminpass')
        self.module = Module.objects.create(name='Programmation Web', code='WEB301')

    def build_file(self, rows, headers=None):
        if headers is None:
            headers = ['Prénom', 'Nom', 'Email', 'ID Enseignant', 'Département']
        return build_roster_file(rows, headers=headers)

    def test_import_creates_teachers_and_assignment(self):
        from teacher.models import TeacherProfile, ModuleAssignment

        roster = self.build_file([
            ['Karim', 'Alaoui', 'karim.alaoui@example.com', 'T1001', 'Informatique'],
            ['Salma', 'Bennis', 'salma.bennis@example.com', '', 'Génie Civil'],
        ])
        response = self.client.post(reverse('administrator:import_teachers'), {
            'excel_file': roster,
            'module_id': self.module.id,
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['created']), 2)
        self.assertEqual(response.context['errors'], [])

        teacher = TeacherProfile.objects.get(teacher_id='T1001')
        self.assertEqual(teacher.user.email, 'karim.alaoui@example.com')
        self.assertTrue(ModuleAssignment.objects.filter(
            teacher=teacher, module=self.module, is_active=True
        ).exists())

        # teacher_id was optional for the second row
        self.assertTrue(TeacherProfile.objects.filter(
            user__email='salma.bennis@example.com', teacher_id__isnull=True
        ).exists())

        # Generated password logs the teacher in
        password = response.context['created'][0]['password']
        username = response.context['created'][0]['username']
        self.client.logout()
        self.assertTrue(self.client.login(username=username, password=password))

    def test_import_skips_duplicates(self):
        from teacher.models import TeacherProfile

        existing = User.objects.create_user(
            username='karim', email='karim.alaoui@example.com', password='x'
        )
        TeacherProfile.objects.create(user=existing, teacher_id='T9999', department='Info')

        roster = self.build_file([
            ['Karim', 'Alaoui', 'karim.alaoui@example.com', 'T1001', 'Informatique'],  # dup email
            ['Nadia', 'Cherkaoui', 'nadia.cherkaoui@example.com', 'T9999', 'Info'],     # dup teacher id
            ['Omar', 'Drissi', 'omar.drissi@example.com', '', 'Info'],                  # valid
        ])
        response = self.client.post(reverse('administrator:import_teachers'), {'excel_file': roster})

        self.assertEqual(len(response.context['created']), 1)
        self.assertEqual(len(response.context['skipped']), 2)

    def test_template_download(self):
        response = self.client.get(reverse('administrator:teacher_import_template'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])


class ToggleUserActiveTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(username='admin', password='adminpass')
        AdminProfile.objects.create(user=self.admin_user, admin_id='A001')

        student_user = User.objects.create_user(username='etudiant', password='pass')
        self.student = StudentProfile.objects.create(
            user=student_user, student_id='E001', year_of_study=3, department='Info'
        )

        self.client.login(username='admin', password='adminpass')

    def toggle(self, user):
        return self.client.post(reverse('administrator:toggle_user_active', args=[user.id]))

    def test_deactivate_and_reactivate_student(self):
        response = self.toggle(self.student.user)
        self.assertEqual(response.status_code, 302)
        self.student.user.refresh_from_db()
        self.assertFalse(self.student.user.is_active)

        # Deactivated account can no longer log in
        self.assertFalse(self.client.login(username='etudiant', password='pass'))

        self.client.login(username='admin', password='adminpass')
        self.toggle(self.student.user)
        self.student.user.refresh_from_db()
        self.assertTrue(self.student.user.is_active)

    def test_cannot_deactivate_self(self):
        self.toggle(self.admin_user)
        self.admin_user.refresh_from_db()
        self.assertTrue(self.admin_user.is_active)

    def test_cannot_deactivate_other_admin(self):
        other_admin = User.objects.create_user(username='admin2', password='pass')
        AdminProfile.objects.create(user=other_admin, admin_id='A002')

        self.toggle(other_admin)
        other_admin.refresh_from_db()
        self.assertTrue(other_admin.is_active)

    def test_requires_admin(self):
        self.client.logout()
        self.client.login(username='etudiant', password='pass')
        response = self.toggle(self.admin_user)
        self.assertEqual(response.status_code, 302)  # bounced to login


class AdminExportTests(TestCase):
    def setUp(self):
        from datetime import date
        from student.models import Project
        from teacher.models import TeacherProfile

        admin_user = User.objects.create_user(username='admin', password='adminpass')
        AdminProfile.objects.create(user=admin_user, admin_id='A001')

        student_user = User.objects.create_user(
            username='etudiant', password='pass', first_name='Amina', last_name='Benali'
        )
        self.student = StudentProfile.objects.create(
            user=student_user, student_id='E001', year_of_study=3, department='Info'
        )
        teacher_user = User.objects.create_user(
            username='prof', password='pass', first_name='Karim', last_name='Alaoui'
        )
        TeacherProfile.objects.create(user=teacher_user, teacher_id='T001', department='Info')

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')
        self.project = Project.objects.create(
            title='Projet Export', description='d', project_type='module',
            student=self.student, module=self.module,
            start_date=date(2026, 1, 1), status='validated',
        )

        self.client.login(username='admin', password='adminpass')

    def read_workbook(self, response):
        from openpyxl import load_workbook
        return load_workbook(BytesIO(response.content))

    def test_export_projects_excel(self):
        response = self.client.get(reverse('administrator:export_projects'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

        sheet = self.read_workbook(response).active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], 'Titre')
        titles = [row[0] for row in rows[1:]]
        self.assertIn('Projet Export', titles)

    def test_export_projects_respects_status_filter(self):
        response = self.client.get(reverse('administrator:export_projects'), {'status': 'submitted'})
        sheet = self.read_workbook(response).active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)  # header only, no submitted projects

    def test_export_users_excel_has_both_sheets(self):
        response = self.client.get(reverse('administrator:export_users'))
        self.assertEqual(response.status_code, 200)
        workbook = self.read_workbook(response)
        self.assertIn('Étudiants', workbook.sheetnames)
        self.assertIn('Enseignants', workbook.sheetnames)

        student_rows = list(workbook['Étudiants'].iter_rows(values_only=True))
        self.assertIn('Benali', [row[1] for row in student_rows[1:]])
        teacher_rows = list(workbook['Enseignants'].iter_rows(values_only=True))
        self.assertIn('Alaoui', [row[1] for row in teacher_rows[1:]])

    def test_export_statistics_pdf(self):
        response = self.client.get(reverse('administrator:export_statistics'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_exports_require_admin(self):
        self.client.logout()
        self.client.login(username='etudiant', password='pass')
        response = self.client.get(reverse('administrator:export_projects'))
        self.assertEqual(response.status_code, 302)  # bounced to login

    def test_exports_page_renders(self):
        response = self.client.get(reverse('administrator:exports'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Exports')
