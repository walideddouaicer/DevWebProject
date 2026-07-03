# Excel / PDF export helpers (ROADMAP #7).
from django.http import HttpResponse
from django.utils import timezone

EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def excel_response(workbook, filename):
    """Serialize an openpyxl workbook into a download response."""
    response = HttpResponse(content_type=EXCEL_CONTENT_TYPE)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _style_header(sheet):
    """Bold white-on-blue header row + freeze panes."""
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4299E1', end_color='4299E1', fill_type='solid')
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    sheet.freeze_panes = 'A2'


def _autosize(sheet, max_width=50):
    for column_cells in sheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        letter = column_cells[0].column_letter
        sheet.column_dimensions[letter].width = min(length + 3, max_width)


def build_projects_workbook(projects):
    """Excel export of a project queryset."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Projets'
    sheet.append([
        'Titre', 'Étudiant', 'ID Étudiant', 'Collaborateurs', 'Module',
        'Type', 'Statut', 'Note /20', 'Public', 'Créé le', 'Date début', 'Date fin',
    ])

    for project in projects:
        evaluation = getattr(project, 'evaluation', None)
        collaborators = ", ".join(
            c.get_full_name() for c in project.collaborators.all()
        )
        sheet.append([
            project.title,
            project.student.get_full_name(),
            project.student.student_id,
            collaborators,
            project.module.code if project.module else '',
            project.get_project_type_display(),
            project.get_status_display(),
            float(evaluation.score) if evaluation else '',
            'Oui' if project.is_public else 'Non',
            project.created_at.strftime('%d/%m/%Y'),
            project.start_date.strftime('%d/%m/%Y') if project.start_date else '',
            project.end_date.strftime('%d/%m/%Y') if project.end_date else '',
        ])

    _style_header(sheet)
    _autosize(sheet)
    return workbook


def build_users_workbook(students, teachers):
    """Excel export of all users: one sheet per role."""
    from openpyxl import Workbook

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = 'Étudiants'
    sheet.append(['Prénom', 'Nom', 'Nom d\'utilisateur', 'Email', 'ID Étudiant',
                  'Année', 'Département', 'Inscrit le'])
    for s in students:
        sheet.append([
            s.user.first_name, s.user.last_name, s.user.username, s.user.email,
            s.student_id, s.get_year_of_study_display(), s.department,
            s.user.date_joined.strftime('%d/%m/%Y'),
        ])
    _style_header(sheet)
    _autosize(sheet)

    sheet2 = workbook.create_sheet('Enseignants')
    sheet2.append(['Prénom', 'Nom', 'Nom d\'utilisateur', 'Email', 'ID Enseignant',
                   'Département', 'Inscrit le'])
    for t in teachers:
        sheet2.append([
            t.user.first_name, t.user.last_name, t.user.username, t.user.email,
            t.teacher_id or '', t.department,
            t.user.date_joined.strftime('%d/%m/%Y'),
        ])
    _style_header(sheet2)
    _autosize(sheet2)

    return workbook


def build_module_roster_workbook(module, enrollments):
    """Excel export of one module's enrolled students."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = module.code[:31]
    sheet.append(['Prénom', 'Nom', 'ID Étudiant', 'Email', 'Année', 'Département', 'Inscrit le'])
    for enrollment in enrollments:
        s = enrollment.student
        sheet.append([
            s.user.first_name, s.user.last_name, s.student_id, s.user.email,
            s.get_year_of_study_display(), s.department,
            enrollment.enrolled_at.strftime('%d/%m/%Y'),
        ])
    _style_header(sheet)
    _autosize(sheet)
    return workbook


def build_assignment_results_workbook(assignment, projects):
    """Excel export of an assignment's teams, statuses and grades."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Résultats'
    sheet.append([
        'Projet', 'Chef d\'équipe', 'ID Étudiant', 'Équipe', 'Taille',
        'Option choisie', 'Statut', 'Soumis le', 'Note /20', 'Mention', 'Appréciation',
    ])

    for project in projects:
        evaluation = getattr(project, 'evaluation', None)
        team = ", ".join(m.get_full_name() for m in project.get_team_members())
        sheet.append([
            project.title,
            project.student.get_full_name(),
            project.student.student_id,
            team,
            project.get_team_size(),
            project.selected_option.title if project.selected_option else '',
            project.get_status_display(),
            project.assignment_submitted_at.strftime('%d/%m/%Y %H:%M') if project.assignment_submitted_at else '',
            float(evaluation.score) if evaluation else '',
            evaluation.get_mention() if evaluation else '',
            evaluation.comments if evaluation else '',
        ])

    _style_header(sheet)
    _autosize(sheet)
    return workbook


def build_statistics_pdf():
    """PDF summary report of platform statistics (reportlab)."""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    from student.models import Project, StudentProfile
    from teacher.models import Module, ModuleEnrollment, TeacherProfile

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], textColor=colors.HexColor('#2d3748'))

    def stat_table(rows):
        table = Table(rows, colWidths=[9 * cm, 5 * cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4299e1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        return table

    elements = [
        Paragraph("ENSA Project Manager — Rapport statistique", title_style),
        Paragraph(f"Généré le {timezone.localtime(timezone.now()).strftime('%d/%m/%Y à %H:%M')}", styles['Normal']),
        Spacer(1, 16),
    ]

    # Global counts
    elements.append(Paragraph("Vue d'ensemble", styles['Heading2']))
    elements.append(stat_table([
        ['Indicateur', 'Valeur'],
        ['Étudiants', StudentProfile.objects.count()],
        ['Enseignants', TeacherProfile.objects.count()],
        ['Modules actifs', Module.objects.filter(is_active=True).count()],
        ['Inscriptions aux modules', ModuleEnrollment.objects.filter(is_active=True).count()],
        ['Projets', Project.objects.count()],
        ['Projets publics', Project.objects.filter(is_public=True).count()],
    ]))
    elements.append(Spacer(1, 16))

    # Status breakdown
    elements.append(Paragraph("Projets par statut", styles['Heading2']))
    status_rows = [['Statut', 'Nombre']]
    for status_key, status_name in Project.STATUS_CHOICES:
        status_rows.append([status_name, Project.objects.filter(status=status_key).count()])
    elements.append(stat_table(status_rows))
    elements.append(Spacer(1, 16))

    # Type breakdown
    elements.append(Paragraph("Projets par type", styles['Heading2']))
    type_rows = [['Type', 'Nombre']]
    for type_key, type_name in Project.PROJECT_TYPES:
        type_rows.append([type_name, Project.objects.filter(project_type=type_key).count()])
    elements.append(stat_table(type_rows))
    elements.append(Spacer(1, 16))

    # Top modules
    elements.append(Paragraph("Modules les plus actifs", styles['Heading2']))
    module_rows = [['Module', 'Étudiants / Projets']]
    for module in Module.objects.filter(is_active=True).order_by('code')[:15]:
        module_rows.append([
            f"{module.code} — {module.name}",
            f"{module.get_enrolled_students_count()} étudiants · {module.projects.count()} projets",
        ])
    elements.append(stat_table(module_rows))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="statistiques_ensa.pdf"'
    return response
