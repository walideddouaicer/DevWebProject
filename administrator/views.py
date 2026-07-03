# administrator/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from django.contrib.auth.models import User
from datetime import datetime, timedelta

# Import models from other apps
from .models import AdminProfile
from teacher.models import TeacherProfile, Module, ModuleAssignment, ModuleEnrollment
from student.models import StudentProfile, Project, ProjectActivity
from accounts.models import PendingRegistration

# Authentication decorator for administrators
def admin_required(view_func):
    """Decorator to ensure user is an administrator"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        
        try:
            AdminProfile.objects.get(user=request.user)
        except AdminProfile.DoesNotExist:
            messages.error(request, "Vous n'avez pas les droits d'administrateur.")
            return redirect('login')
        
        return view_func(request, *args, **kwargs)
    return wrapper

@admin_required
def dashboard(request):
    """Administrator dashboard with key statistics - FIXED VERSION"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Basic statistics - these work correctly
    total_projects = Project.objects.count()
    total_students = StudentProfile.objects.count()
    total_teachers = TeacherProfile.objects.count()
    total_modules = Module.objects.filter(is_active=True).count()
    
    # Project status breakdown - calculate manually for accuracy
    projects_in_progress = Project.objects.filter(status='in_progress').count()
    projects_submitted = Project.objects.filter(status='submitted').count()
    projects_validated = Project.objects.filter(status='validated').count()
    projects_rejected = Project.objects.filter(status='rejected').count()
    
    # Recent activity (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_projects = Project.objects.filter(created_at__gte=thirty_days_ago).count()
    recent_enrollments = ModuleEnrollment.objects.filter(enrolled_at__gte=thirty_days_ago).count()
    
    # FIXED: Module statistics with manual counting (this is what works!)
    modules_with_stats = []
    
    for module in Module.objects.filter(is_active=True).order_by('code'):
        # Manual count for each module - this is the pattern that works
        student_count = ModuleEnrollment.objects.filter(
            module=module, 
            is_active=True
        ).count()
        
        teacher_count = ModuleAssignment.objects.filter(
            module=module, 
            is_active=True
        ).count()
        
        project_count = Project.objects.filter(module=module).count()
        
        # Create a module dict with stats
        module_data = {
            'id': module.id,
            'code': module.code,
            'name': module.name,
            'academic_year': module.academic_year,
            'semester': module.semester,
            'get_semester_display': module.get_semester_display(),
            'student_count': student_count,
            'teacher_count': teacher_count,
            'project_count': project_count,
        }
        
        modules_with_stats.append(module_data)
    
    # Sort by student count (descending) and take top 5
    modules_with_stats.sort(key=lambda x: x['student_count'], reverse=True)
    modules_with_stats = modules_with_stats[:5]
    
    # Recent projects for activity feed
    latest_projects = Project.objects.select_related(
        'student__user', 'module'
    ).order_by('-updated_at')[:10]
    
    # Pending registrations count
    pending_registrations_count = PendingRegistration.objects.filter(is_approved=False).count()
    
    context = {
        'admin': admin,
        'total_projects': total_projects,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_modules': total_modules,
        'projects_in_progress': projects_in_progress,
        'projects_submitted': projects_submitted,
        'projects_validated': projects_validated,
        'projects_rejected': projects_rejected,
        'recent_projects': recent_projects,
        'recent_enrollments': recent_enrollments,
        'modules_with_stats': modules_with_stats,
        'latest_projects': latest_projects,
        'pending_registrations_count': pending_registrations_count,
    }
    
    return render(request, 'administrator/dashboard.html', context)

@admin_required
def projects_list(request):
    """List all projects with filtering capabilities - FIXED VERSION"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Start with all projects
    projects = Project.objects.select_related(
        'student__user', 'module'
    ).prefetch_related('collaborators__user')
    
    # Apply filters (existing filter logic...)
    search_query = request.GET.get('search', '')
    if search_query:
        projects = projects.filter(
            Q(title__icontains=search_query) |
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__user__username__icontains=search_query)
        )
    
    # Filter by module
    module_id = request.GET.get('module')
    if module_id:
        projects = projects.filter(module_id=module_id)
    
    # Filter by teacher (through module assignments)
    teacher_id = request.GET.get('teacher')
    if teacher_id:
        teacher_modules = ModuleAssignment.objects.filter(
            teacher_id=teacher_id, is_active=True
        ).values_list('module_id', flat=True)
        projects = projects.filter(module_id__in=teacher_modules)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        projects = projects.filter(status=status)
    
    # Filter by project type
    project_type = request.GET.get('project_type')
    if project_type:
        projects = projects.filter(project_type=project_type)
    
    # Filter by academic year
    academic_year = request.GET.get('academic_year')
    if academic_year:
        projects = projects.filter(module__academic_year=academic_year)
    
    # Filter by date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        projects = projects.filter(created_at__gte=date_from)
    if date_to:
        projects = projects.filter(created_at__lte=date_to)
    
    # Order by latest first
    projects = projects.order_by('-updated_at')
    
    # FIXED: Calculate statistics manually
    total_projects = projects.count()
    submitted_count = projects.filter(status='submitted').count()
    validated_count = projects.filter(status='validated').count()
    rejected_count = projects.filter(status='rejected').count()
    in_progress_count = projects.filter(status='in_progress').count()
    
    # Pagination
    paginator = Paginator(projects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    modules = Module.objects.filter(is_active=True).order_by('code')
    teachers = TeacherProfile.objects.select_related('user').order_by('user__first_name')
    academic_years = Module.objects.values_list('academic_year', flat=True).distinct().order_by('academic_year')
    
    context = {
        'admin': admin,
        'page_obj': page_obj,
        'modules': modules,
        'teachers': teachers,
        'academic_years': academic_years,
        'search_query': search_query,
        'project_types': Project.PROJECT_TYPES,
        'status_choices': Project.STATUS_CHOICES,
        # FIXED: Add manually calculated counts
        'total_projects': total_projects,
        'submitted_count': submitted_count,
        'validated_count': validated_count,
        'rejected_count': rejected_count,
        'in_progress_count': in_progress_count,
        'selected_module': module_id,
        'selected_teacher': teacher_id,
        'selected_status': status,
        'selected_project_type': project_type,
        'selected_academic_year': academic_year,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'administrator/projects_list.html', context)

@admin_required
def project_detail(request, project_id):
    """View detailed information about a specific project (read-only)"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    project = get_object_or_404(
        Project.objects.select_related('student__user', 'module'),
        id=project_id
    )
    
    # Get project details
    deliverables = project.deliverables.all().order_by('-upload_date')
    milestones = project.milestones.all().order_by('due_date')
    comments = project.comments.all().select_related('author').order_by('-created_at')
    activities = project.activities.all().select_related('user').order_by('-created_at')
    collaborators = project.collaborators.all().select_related('user')
    
    # Get assigned teachers for this project's module
    assigned_teachers = []
    if project.module:
        assignments = ModuleAssignment.objects.filter(
            module=project.module,
            is_active=True
        ).select_related('teacher__user')
        assigned_teachers = [assignment.teacher for assignment in assignments]
    
    context = {
        'admin': admin,
        'project': project,
        'deliverables': deliverables,
        'milestones': milestones,
        'comments': comments,
        'activities': activities,
        'collaborators': collaborators,
        'assigned_teachers': assigned_teachers,
    }
    
    return render(request, 'administrator/project_detail.html', context)

@admin_required
def modules_list(request):
    """List all modules with management options - FIXED VERSION"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Get base queryset
    modules = Module.objects.all().order_by('-created_at')
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        modules = modules.filter(
            Q(code__icontains=search_query) |
            Q(name__icontains=search_query)
        )
    
    academic_year = request.GET.get('academic_year')
    if academic_year:
        modules = modules.filter(academic_year=academic_year)
    
    semester = request.GET.get('semester')
    if semester:
        modules = modules.filter(semester=semester)
    
    # FIXED: Calculate totals manually for accuracy
    total_modules = modules.count()
    active_modules = modules.filter(is_active=True).count()
    
    # Calculate total students and teachers across all modules
    total_students_enrolled = 0
    total_teachers_assigned = 0
    
    for module in modules:
        total_students_enrolled += ModuleEnrollment.objects.filter(
            module=module, is_active=True
        ).count()
        total_teachers_assigned += ModuleAssignment.objects.filter(
            module=module, is_active=True
        ).count()
    
    # FIXED: Add manual counts to each module for template use
    modules_with_counts = []
    for module in modules:
        student_count = ModuleEnrollment.objects.filter(
            module=module, is_active=True
        ).count()
        teacher_count = ModuleAssignment.objects.filter(
            module=module, is_active=True
        ).count()
        project_count = Project.objects.filter(module=module).count()
        
        # Add counts as attributes
        module.student_count = student_count
        module.teacher_count = teacher_count  
        module.project_count = project_count
        
        modules_with_counts.append(module)
    
    # Get filter options
    academic_years = Module.objects.values_list('academic_year', flat=True).distinct().order_by('academic_year')
    
    context = {
        'admin': admin,
        'modules': modules_with_counts,  # Now has count attributes
        'academic_years': academic_years,
        'search_query': search_query,
        'selected_academic_year': academic_year,
        'selected_semester': semester,
        'semester_choices': Module._meta.get_field('semester').choices,
        # FIXED: Add totals calculated manually
        'total_modules': total_modules,
        'active_modules': active_modules,
        'total_students_enrolled': total_students_enrolled,
        'total_teachers_assigned': total_teachers_assigned,
    }
    
    return render(request, 'administrator/modules_list.html', context)

@admin_required
def module_create(request):
    """Create a new module"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    if request.method == 'POST':
        # Get form data
        name = request.POST.get('name')
        code = request.POST.get('code')
        description = request.POST.get('description', '')
        academic_year = request.POST.get('academic_year')
        semester = request.POST.get('semester')
        
        # Validate required fields
        if not all([name, code, academic_year, semester]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
            return render(request, 'administrator/module_create.html', {'admin': admin})
        
        # Check if module code already exists
        if Module.objects.filter(code=code).exists():
            messages.error(request, f"Un module avec le code '{code}' existe déjà.")
            return render(request, 'administrator/module_create.html', {'admin': admin})
        
        # Create the module
        try:
            module = Module.objects.create(
                name=name,
                code=code.upper(),
                description=description,
                academic_year=academic_year,
                semester=semester,
                is_active=True
            )
            
            messages.success(request, f"Module '{module.code} - {module.name}' créé avec succès.")
            return redirect('administrator:module_detail', module_id=module.id)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création du module: {str(e)}")
    
    context = {
        'admin': admin,
        'semester_choices': Module._meta.get_field('semester').choices,
    }
    
    return render(request, 'administrator/module_create.html', context)

@admin_required
def module_detail(request, module_id):
    """View detailed information about a module"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    module = get_object_or_404(Module, id=module_id)
    
    # Get module statistics
    enrollments = ModuleEnrollment.objects.filter(
        module=module, is_active=True
    ).select_related('student__user')
    
    assignments = ModuleAssignment.objects.filter(
        module=module, is_active=True
    ).select_related('teacher__user')
    
    module_projects = Project.objects.filter(module=module).select_related('student__user')
    
    context = {
        'admin': admin,
        'module': module,
        'enrollments': enrollments,
        'assignments': assignments,
        'module_projects': module_projects,
        'total_students': enrollments.count(),
        'total_teachers': assignments.count(),
        'total_projects': module_projects.count(),
    }
    
    return render(request, 'administrator/module_detail.html', context)

@admin_required
def module_edit(request, module_id):
    """Edit module information"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    module = get_object_or_404(Module, id=module_id)
    
    if request.method == 'POST':
        # Update module fields
        module.name = request.POST.get('name', module.name)
        module.description = request.POST.get('description', module.description)
        module.academic_year = request.POST.get('academic_year', module.academic_year)
        module.semester = request.POST.get('semester', module.semester)
        module.is_active = request.POST.get('is_active') == 'on'
        
        try:
            module.save()
            messages.success(request, f"Module '{module.code}' mis à jour avec succès.")
            return redirect('administrator:module_detail', module_id=module.id)
        except Exception as e:
            messages.error(request, f"Erreur lors de la mise à jour: {str(e)}")
    
    context = {
        'admin': admin,
        'module': module,
        'semester_choices': Module._meta.get_field('semester').choices,
    }
    
    return render(request, 'administrator/module_edit.html', context)

@admin_required
def module_delete(request, module_id):
    """Delete a module (soft delete - set is_active to False)"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    module = get_object_or_404(Module, id=module_id)
    
    if request.method == 'POST':
        # Soft delete - set is_active to False
        module.is_active = False
        module.save()
        
        messages.success(request, f"Module '{module.code}' désactivé avec succès.")
        return redirect('administrator:modules_list')
    
    # Get statistics for confirmation
    student_count = ModuleEnrollment.objects.filter(module=module, is_active=True).count()
    teacher_count = ModuleAssignment.objects.filter(module=module, is_active=True).count()
    project_count = Project.objects.filter(module=module).count()
    
    context = {
        'admin': admin,
        'module': module,
        'student_count': student_count,
        'teacher_count': teacher_count,
        'project_count': project_count,
    }
    
    return render(request, 'administrator/module_delete.html', context)

@admin_required
def assignments_management(request):
    """Manage teacher-module assignments - FIXED VERSION"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Get all active assignments
    assignments = ModuleAssignment.objects.filter(
        is_active=True
    ).select_related('teacher__user', 'module').order_by('module__code')
    
    # Get teachers and modules for new assignments
    teachers = TeacherProfile.objects.select_related('user').order_by('user__first_name')
    modules = Module.objects.filter(is_active=True).order_by('code')
    
    # FIXED: Calculate unassigned modules manually
    assigned_module_ids = assignments.values_list('module_id', flat=True)
    unassigned_modules = modules.exclude(id__in=assigned_module_ids)
    
    # FIXED: Calculate statistics manually for accuracy
    total_assignments = assignments.count()
    total_teachers = teachers.count()
    total_modules = modules.count()
    total_unassigned = unassigned_modules.count()
    
    # FIXED: Add workload counts to teachers for template
    teachers_with_workload = []
    for teacher in teachers:
        teacher_assignments = ModuleAssignment.objects.filter(
            teacher=teacher, is_active=True
        ).count()
        
        # Calculate total students taught by this teacher
        teacher_modules = ModuleAssignment.objects.filter(
            teacher=teacher, is_active=True
        ).values_list('module_id', flat=True)
        
        total_students = 0
        for module_id in teacher_modules:
            total_students += ModuleEnrollment.objects.filter(
                module_id=module_id, is_active=True
            ).count()
        
        # Add attributes for template
        teacher.assignment_count = teacher_assignments
        teacher.student_count = total_students
        
        teachers_with_workload.append(teacher)
    
    context = {
        'admin': admin,
        'assignments': assignments,
        'teachers': teachers_with_workload,  # Now includes workload data
        'modules': modules,
        'unassigned_modules': unassigned_modules,
        # FIXED: Add manual counts
        'total_assignments': total_assignments,
        'total_teachers': total_teachers,
        'total_modules': total_modules,
        'total_unassigned': total_unassigned,
    }
    
    return render(request, 'administrator/assignments_management.html', context)

@admin_required
def assign_teacher_to_module(request):
    """Assign a teacher to a module"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        module_id = request.POST.get('module_id')
        
        if not teacher_id or not module_id:
            messages.error(request, "Veuillez sélectionner un enseignant et un module.")
            return redirect('administrator:assignments_management')
        
        try:
            teacher = TeacherProfile.objects.get(id=teacher_id)
            module = Module.objects.get(id=module_id)
            
            # Check if assignment already exists
            existing_assignment = ModuleAssignment.objects.filter(
                teacher=teacher, module=module
            ).first()
            
            if existing_assignment:
                if existing_assignment.is_active:
                    messages.info(request, f"{teacher.user.get_full_name()} est déjà assigné au module {module.code}.")
                else:
                    # Reactivate existing assignment
                    existing_assignment.is_active = True
                    existing_assignment.assigned_by = request.user
                    existing_assignment.save()
                    messages.success(request, f"{teacher.user.get_full_name()} a été réassigné au module {module.code}.")
            else:
                # Create new assignment
                ModuleAssignment.objects.create(
                    teacher=teacher,
                    module=module,
                    assigned_by=request.user,
                    is_active=True
                )
                messages.success(request, f"{teacher.user.get_full_name()} a été assigné au module {module.code}.")
                
        except (TeacherProfile.DoesNotExist, Module.DoesNotExist) as e:
            messages.error(request, "Enseignant ou module introuvable.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'assignation: {str(e)}")
    
    return redirect('administrator:assignments_management')

@admin_required
def remove_assignment(request, assignment_id):
    """Remove a teacher-module assignment"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    try:
        assignment = ModuleAssignment.objects.get(id=assignment_id, is_active=True)
        assignment.is_active = False
        assignment.save()
        
        messages.success(request, f"Assignation de {assignment.teacher.user.get_full_name()} au module {assignment.module.code} supprimée.")
    except ModuleAssignment.DoesNotExist:
        messages.error(request, "Assignation introuvable.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression: {str(e)}")
    
    return redirect('administrator:assignments_management')

@admin_required
def users_overview(request):
    """Overview of all users in the system"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # User statistics
    total_students = StudentProfile.objects.count()
    total_teachers = TeacherProfile.objects.count()
    total_admins = AdminProfile.objects.count()
    
    # Recent registrations (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_students = StudentProfile.objects.filter(user__date_joined__gte=thirty_days_ago).count()
    recent_teachers = TeacherProfile.objects.filter(user__date_joined__gte=thirty_days_ago).count()
    
    context = {
        'admin': admin,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_admins': total_admins,
        'recent_students': recent_students,
        'recent_teachers': recent_teachers,
    }
    
    return render(request, 'administrator/users_overview.html', context)

@admin_required
def students_list(request):
    """List all students"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    students = StudentProfile.objects.select_related('user').order_by('user__first_name')
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(student_id__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(students, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'admin': admin,
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'administrator/students_list.html', context)

@admin_required
def teachers_list(request):
    """List all teachers"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    teachers = TeacherProfile.objects.select_related('user').order_by('user__first_name')
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        teachers = teachers.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(teacher_id__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(teachers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'admin': admin,
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'administrator/teachers_list.html', context)

@admin_required
def statistics(request):
    """Detailed statistics page"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Project statistics by type
    project_type_stats = []
    for type_key, type_name in Project.PROJECT_TYPES:
        count = Project.objects.filter(project_type=type_key).count()
        project_type_stats.append({'type': type_name, 'count': count})
    
    # Project statistics by status
    status_stats = []
    for status_key, status_name in Project.STATUS_CHOICES:
        count = Project.objects.filter(status=status_key).count()
        status_stats.append({'status': status_name, 'count': count})
    
    # Module statistics
    module_stats = Module.objects.filter(is_active=True).annotate(
        student_count=Count('enrollments', filter=Q(enrollments__is_active=True)),
        project_count=Count('projects')
    ).order_by('-student_count')
    
    # Teacher workload
    teacher_workload = TeacherProfile.objects.annotate(
        module_count=Count('module_assignments', filter=Q(module_assignments__is_active=True)),
        student_count=Count('module_assignments__module__enrollments', 
                          filter=Q(module_assignments__is_active=True, 
                                  module_assignments__module__enrollments__is_active=True))
    ).order_by('-module_count')
    
    context = {
        'admin': admin,
        'project_type_stats': project_type_stats,
        'status_stats': status_stats,
        'module_stats': module_stats,
        'teacher_workload': teacher_workload,
    }
    
    return render(request, 'administrator/statistics.html', context)

@admin_required
def remove_student_from_module(request, module_id, student_id):
    """Remove a student from a module"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    try:
        module = Module.objects.get(id=module_id)
        student = StudentProfile.objects.get(id=student_id)
        
        # Find the enrollment
        enrollment = ModuleEnrollment.objects.get(
            module=module,
            student=student,
            is_active=True
        )
        
        # Deactivate the enrollment (soft delete)
        enrollment.is_active = False
        enrollment.save()
        
        if request.method == 'POST':
            # For AJAX requests
            return JsonResponse({
                'success': True, 
                'message': f'{student.user.get_full_name() or student.user.username} a été désinscrit du module {module.code}.'
            })
        else:
            # For regular requests
            messages.success(request, f'{student.user.get_full_name() or student.user.username} a été désinscrit du module {module.code}.')
            return redirect('administrator:module_detail', module_id=module.id)
            
    except ModuleEnrollment.DoesNotExist:
        error_msg = "Cet étudiant n'est pas inscrit à ce module."
        if request.method == 'POST':
            return JsonResponse({'success': False, 'message': error_msg})
        else:
            messages.error(request, error_msg)
            return redirect('administrator:module_detail', module_id=module_id)
    except (Module.DoesNotExist, StudentProfile.DoesNotExist):
        error_msg = "Module ou étudiant introuvable."
        if request.method == 'POST':
            return JsonResponse({'success': False, 'message': error_msg})
        else:
            messages.error(request, error_msg)
            return redirect('administrator:modules_list')
    except Exception as e:
        error_msg = f"Erreur lors de la désinscription: {str(e)}"
        if request.method == 'POST':
            return JsonResponse({'success': False, 'message': error_msg})
        else:
            messages.error(request, error_msg)
            return redirect('administrator:module_detail', module_id=module_id)

@admin_required
def exports(request):
    """Export options page"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    context = {
        'admin': admin,
    }
    
    return render(request, 'administrator/exports.html', context)

@admin_required
def export_projects(request):
    """Export all projects to Excel (honours the projects_list filters)"""
    from .exports import build_projects_workbook, excel_response

    projects = Project.objects.select_related(
        'student__user', 'module'
    ).prefetch_related('collaborators__user').order_by('-created_at')

    # Reuse the same GET filters as the projects list
    status = request.GET.get('status')
    if status:
        projects = projects.filter(status=status)
    module_id = request.GET.get('module')
    if module_id:
        projects = projects.filter(module_id=module_id)
    project_type = request.GET.get('project_type')
    if project_type:
        projects = projects.filter(project_type=project_type)

    workbook = build_projects_workbook(projects)
    filename = f"projets_ensa_{timezone.now().strftime('%Y%m%d')}.xlsx"
    return excel_response(workbook, filename)

@admin_required
def export_statistics(request):
    """Export platform statistics as a PDF report"""
    from .exports import build_statistics_pdf
    return build_statistics_pdf()

@admin_required
def pending_registrations(request):
    """View pending user registrations"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    pending = PendingRegistration.objects.filter(
        is_approved=False
    ).order_by('-created_at')
    
    # Get statistics
    total_pending = pending.count()
    student_pending = pending.filter(role='student').count()
    teacher_pending = pending.filter(role='teacher').count()
    
    context = {
        'admin': admin,
        'pending_registrations': pending,
        'total_pending': total_pending,
        'student_pending': student_pending,
        'teacher_pending': teacher_pending,
    }
    
    return render(request, 'administrator/pending_registrations.html', context)

@admin_required  
def approve_registration(request, registration_id):
    """Approve a pending registration"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    registration = get_object_or_404(PendingRegistration, id=registration_id, is_approved=False)
    
    if request.method == 'POST':
        from django.db import transaction
        from accounts.utils import send_approval_email

        try:
            # All-or-nothing: if profile creation fails (e.g. duplicate
            # student_id), no orphan User account is left behind.
            with transaction.atomic():
                user = User.objects.create(
                    username=registration.username,
                    email=registration.email,
                    first_name=registration.first_name,
                    last_name=registration.last_name,
                    is_active=True
                )
                # Set the already hashed password directly
                user.password = registration.password
                user.save()

                # Create appropriate profile
                if registration.role == 'student':
                    StudentProfile.objects.create(
                        user=user,
                        student_id=registration.student_id,
                        year_of_study=registration.year_of_study,
                        department=registration.department
                    )
                    profile_type = "étudiant"
                elif registration.role == 'teacher':
                    TeacherProfile.objects.create(
                        user=user,
                        teacher_id=registration.teacher_id or '',
                        department=registration.department
                    )
                    profile_type = "enseignant"
                else:
                    profile_type = registration.get_role_display().lower()

                # Mark as approved
                registration.is_approved = True
                registration.approved_by = request.user
                registration.approved_at = timezone.now()
                registration.save()

            messages.success(request, f"Compte {profile_type} approuvé pour {registration.first_name} {registration.last_name}")

            # Notify the user by email (console backend in dev)
            send_approval_email(registration)

        except Exception as e:
            messages.error(request, f"Erreur lors de l'approbation: {str(e)}")
    
    return redirect('administrator:pending_registrations')

@admin_required
def reports_moderation(request):
    """Moderation queue for reported public projects (ROADMAP #2)"""
    from student.models import ProjectReport

    admin = get_object_or_404(AdminProfile, user=request.user)

    show = request.GET.get('show', 'pending')  # 'pending', 'hidden', 'all'

    projects = Project.objects.filter(
        reports__isnull=False
    ).distinct().select_related('student__user', 'module').prefetch_related(
        'reports__reporter'
    )

    if show == 'pending':
        projects = projects.filter(reports__is_reviewed=False).distinct()
    elif show == 'hidden':
        projects = projects.filter(is_hidden_by_admin=True)

    projects = projects.order_by('-report_count', '-updated_at')

    # Statistics
    total_reports = ProjectReport.objects.count()
    pending_reports = ProjectReport.objects.filter(is_reviewed=False).count()
    hidden_projects = Project.objects.filter(is_hidden_by_admin=True).count()

    context = {
        'admin': admin,
        'projects': projects,
        'show': show,
        'total_reports': total_reports,
        'pending_reports': pending_reports,
        'hidden_projects': hidden_projects,
    }

    return render(request, 'administrator/reports_moderation.html', context)


@admin_required
def moderate_project(request, project_id):
    """Apply a moderation action to a reported project"""
    from student.models import Notification

    if request.method != 'POST':
        return redirect('administrator:reports_moderation')

    project = get_object_or_404(Project, id=project_id)
    action = request.POST.get('action')
    notes = request.POST.get('admin_notes', '').strip()

    if action == 'hide':
        project.is_hidden_by_admin = True
        project.is_public = False
        project.save(update_fields=['is_hidden_by_admin', 'is_public'])
        project.reports.filter(is_reviewed=False).update(is_reviewed=True, admin_notes=notes)

        Notification.objects.create(
            recipient=project.student,
            project=project,
            notification_type='project_update',
            message=(
                f"Votre projet '{project.title}' a été retiré de la vitrine publique "
                f"suite à des signalements.{' Motif: ' + notes if notes else ''}"
            )
        )
        messages.success(request, f"Projet '{project.title}' masqué de la vitrine publique.")

    elif action == 'unhide':
        project.is_hidden_by_admin = False
        project.is_reported = False
        project.save(update_fields=['is_hidden_by_admin', 'is_reported'])

        Notification.objects.create(
            recipient=project.student,
            project=project,
            notification_type='project_update',
            message=f"Votre projet '{project.title}' est de nouveau autorisé sur la vitrine publique."
        )
        messages.success(request, f"Projet '{project.title}' réautorisé.")

    elif action == 'dismiss':
        project.reports.filter(is_reviewed=False).update(is_reviewed=True, admin_notes=notes)
        project.is_reported = False
        project.save(update_fields=['is_reported'])
        messages.success(request, f"Signalements du projet '{project.title}' classés sans suite.")

    else:
        messages.error(request, "Action de modération invalide.")

    return redirect('administrator:reports_moderation')


@admin_required
def reject_registration(request, registration_id):
    """Reject a pending registration"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    registration = get_object_or_404(PendingRegistration, id=registration_id, is_approved=False)
    
    if request.method == 'POST':
        try:
            name = f"{registration.first_name} {registration.last_name}"
            
            # TODO: Send rejection email before deleting
            registration.delete()
            
            messages.info(request, f"Demande d'inscription de {name} rejetée")
        except Exception as e:
            messages.error(request, f"Erreur lors du rejet: {str(e)}")
    
    return redirect('administrator:pending_registrations')






# Users information fetching


@admin_required
def users_list(request):
    """Unified users list with toggle between students and teachers"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Get user type toggle (default to students)
    user_type = request.GET.get('type', 'students')  # 'students' or 'teachers'
    
    # Apply basic filters
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    year_filter = request.GET.get('year_of_study', '')  # Define this for both user types
    
    if user_type == 'teachers':
        # Get teachers data
        users = TeacherProfile.objects.select_related('user').order_by('user__first_name')
        
        # Apply search filter for teachers
        if search_query:
            users = users.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(user__email__icontains=search_query) |
                Q(teacher_id__icontains=search_query)
            )
        
        # Filter by department
        if department_filter:
            users = users.filter(department__icontains=department_filter)
        
        # Add computed fields for teachers
        users_with_stats = []
        for teacher in users:
            # Calculate assigned modules count
            modules_count = ModuleAssignment.objects.filter(
                teacher=teacher, is_active=True
            ).count()
            
            # Calculate total students taught
            teacher_modules = ModuleAssignment.objects.filter(
                teacher=teacher, is_active=True
            ).values_list('module_id', flat=True)
            
            students_taught = 0
            for module_id in teacher_modules:
                students_taught += ModuleEnrollment.objects.filter(
                    module_id=module_id, is_active=True
                ).count()
            
            # Add computed attributes
            teacher.modules_count = modules_count
            teacher.students_taught = students_taught
            users_with_stats.append(teacher)
        
        users = users_with_stats
        
        # Get all teacher departments for filter
        all_departments = TeacherProfile.objects.values_list('department', flat=True).distinct().order_by('department')
        
    else:  # students
        # Get students data
        users = StudentProfile.objects.select_related('user').order_by('user__first_name')
        
        # Apply search filter for students
        if search_query:
            users = users.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(user__email__icontains=search_query) |
                Q(student_id__icontains=search_query)
            )
        
        # Filter by department
        if department_filter:
            users = users.filter(department__icontains=department_filter)
        
        # Filter by year of study (students only)
        if year_filter:
            users = users.filter(year_of_study=year_filter)
        
        # Add computed fields for students
        users_with_stats = []
        for student in users:
            # Calculate active modules count
            modules_count = ModuleEnrollment.objects.filter(
                student=student, is_active=True
            ).count()
            
            # Calculate projects count
            projects_count = Project.objects.filter(
                Q(student=student) | Q(collaborators=student)
            ).distinct().count()
            
            # Add computed attributes
            student.modules_count = modules_count
            student.projects_count = projects_count
            users_with_stats.append(student)
        
        users = users_with_stats
        
        # Get all student departments for filter
        all_departments = StudentProfile.objects.values_list('department', flat=True).distinct().order_by('department')
    
    # Apply date filters
    if date_from:
        users = [u for u in users if u.user.date_joined.date() >= datetime.strptime(date_from, '%Y-%m-%d').date()]
    if date_to:
        users = [u for u in users if u.user.date_joined.date() <= datetime.strptime(date_to, '%Y-%m-%d').date()]
    
    # Calculate statistics
    total_students = StudentProfile.objects.count()
    total_teachers = TeacherProfile.objects.count()
    
    # Recent registrations (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_students = StudentProfile.objects.filter(user__date_joined__gte=thirty_days_ago).count()
    recent_teachers = TeacherProfile.objects.filter(user__date_joined__gte=thirty_days_ago).count()
    
    # Department breakdown for current user type
    if user_type == 'teachers':
        dept_breakdown = TeacherProfile.objects.values('department').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
    else:
        dept_breakdown = StudentProfile.objects.values('department').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
    
    # Year breakdown (students only)
    year_breakdown = []
    if user_type == 'students':
        year_breakdown = StudentProfile.objects.values('year_of_study').annotate(
            count=Count('id')
        ).order_by('year_of_study')
    
    # Pagination
    paginator = Paginator(users, 25)  # Show 25 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get year choices for students filter
    year_choices = [(3, '3ème Année'), (4, '4ème Année'), (5, '5ème Année')]
    
    context = {
        'admin': admin,
        'page_obj': page_obj,
        'user_type': user_type,
        'search_query': search_query,
        'department_filter': department_filter,
        'date_from': date_from,
        'date_to': date_to,
        'selected_year': year_filter,
        'all_departments': all_departments,
        'year_choices': year_choices,
        
        # Statistics
        'total_students': total_students,
        'total_teachers': total_teachers,
        'recent_students': recent_students,
        'recent_teachers': recent_teachers,
        'current_type_count': len(users),
        'dept_breakdown': dept_breakdown,
        'year_breakdown': year_breakdown,
    }
    
    return render(request, 'administrator/users_list.html', context)

@admin_required
def export_users(request):
    """Export all users to Excel (one sheet per role)"""
    from .exports import build_users_workbook, excel_response

    students = StudentProfile.objects.select_related('user').order_by('user__last_name')
    teachers = TeacherProfile.objects.select_related('user').order_by('user__last_name')

    workbook = build_users_workbook(students, teachers)
    filename = f"utilisateurs_ensa_{timezone.now().strftime('%Y%m%d')}.xlsx"
    return excel_response(workbook, filename)


# Bulk student import from Excel

# Accepted column headers (French and English) mapped to internal field names
STUDENT_IMPORT_HEADER_ALIASES = {
    'prenom': 'first_name', 'prénom': 'first_name', 'first_name': 'first_name',
    'nom': 'last_name', 'last_name': 'last_name',
    'email': 'email', 'e-mail': 'email',
    'id etudiant': 'student_id', 'id étudiant': 'student_id',
    'student_id': 'student_id', 'cne': 'student_id',
    'annee': 'year_of_study', 'année': 'year_of_study',
    'year': 'year_of_study', 'year_of_study': 'year_of_study',
    'departement': 'department', 'département': 'department', 'department': 'department',
}

STUDENT_IMPORT_REQUIRED_FIELDS = ['first_name', 'last_name', 'email', 'student_id', 'year_of_study', 'department']


def _generate_unique_username(email):
    """Build a unique username from the email local part"""
    base = email.split('@')[0].lower().replace(' ', '.')
    username = base
    suffix = 1
    while User.objects.filter(username=username).exists():
        suffix += 1
        username = f"{base}{suffix}"
    return username


@admin_required
def import_students(request):
    """Bulk-create student accounts from an Excel (.xlsx) roster"""
    from django.db import transaction
    from django.utils.crypto import get_random_string
    import openpyxl

    admin = get_object_or_404(AdminProfile, user=request.user)
    modules = Module.objects.filter(is_active=True).order_by('code')

    created = []
    skipped = []
    errors = []
    import_done = False

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        module_id = request.POST.get('module_id', '')

        module = None
        if module_id:
            module = get_object_or_404(Module, id=module_id)

        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx).")
        elif not excel_file.name.lower().endswith('.xlsx'):
            messages.error(request, "Format non supporté. Utilisez un fichier .xlsx (modèle téléchargeable ci-dessous).")
        else:
            try:
                workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet = workbook.active

                rows = sheet.iter_rows(values_only=True)
                header_row = next(rows, None)
                if header_row is None:
                    messages.error(request, "Le fichier est vide.")
                    raise StopIteration

                # Map column index -> field name using header aliases
                column_map = {}
                for index, header in enumerate(header_row):
                    if header is None:
                        continue
                    key = str(header).strip().lower()
                    if key in STUDENT_IMPORT_HEADER_ALIASES:
                        column_map[index] = STUDENT_IMPORT_HEADER_ALIASES[key]

                missing_columns = [f for f in STUDENT_IMPORT_REQUIRED_FIELDS if f not in column_map.values()]
                if missing_columns:
                    messages.error(
                        request,
                        "Colonnes manquantes dans le fichier: Prénom, Nom, Email, ID Étudiant, Année, Département sont requises. "
                        "Téléchargez le modèle pour voir le format attendu."
                    )
                else:
                    for row_number, row in enumerate(rows, start=2):
                        # Skip fully empty rows
                        if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
                            continue

                        data = {}
                        for index, field in column_map.items():
                            value = row[index] if index < len(row) else None
                            data[field] = str(value).strip() if value is not None else ''

                        # Validate required values
                        missing = [f for f in STUDENT_IMPORT_REQUIRED_FIELDS if not data.get(f)]
                        if missing:
                            errors.append({'row': row_number, 'reason': "Valeurs manquantes (toutes les colonnes sont obligatoires)"})
                            continue

                        # Year must be 3, 4 or 5
                        try:
                            year = int(float(data['year_of_study']))
                        except ValueError:
                            year = None
                        if year not in [3, 4, 5]:
                            errors.append({'row': row_number, 'reason': f"Année invalide '{data['year_of_study']}' (3, 4 ou 5 attendu)"})
                            continue

                        email = data['email'].lower()
                        if User.objects.filter(email__iexact=email).exists():
                            skipped.append({'row': row_number, 'identity': email, 'reason': "Email déjà utilisé"})
                            continue
                        if StudentProfile.objects.filter(student_id=data['student_id']).exists():
                            skipped.append({'row': row_number, 'identity': data['student_id'], 'reason': "ID étudiant déjà utilisé"})
                            continue

                        password = get_random_string(10)
                        try:
                            with transaction.atomic():
                                user = User.objects.create_user(
                                    username=_generate_unique_username(email),
                                    email=email,
                                    password=password,
                                    first_name=data['first_name'],
                                    last_name=data['last_name'],
                                    is_active=True,
                                )
                                student = StudentProfile.objects.create(
                                    user=user,
                                    student_id=data['student_id'],
                                    year_of_study=year,
                                    department=data['department'],
                                )
                                if module:
                                    ModuleEnrollment.objects.create(student=student, module=module, is_active=True)
                        except Exception as e:
                            errors.append({'row': row_number, 'reason': f"Erreur lors de la création: {str(e)}"})
                            continue

                        created.append({
                            'name': f"{data['first_name']} {data['last_name']}",
                            'student_id': data['student_id'],
                            'email': email,
                            'username': user.username,
                            'password': password,
                        })

                    import_done = True
                    if created:
                        enrolled_note = f" et inscrits au module {module.code}" if module else ""
                        messages.success(
                            request,
                            f"{len(created)} étudiant(s) créé(s){enrolled_note}. "
                            "Notez les mots de passe ci-dessous: ils ne seront plus affichés."
                        )
                    if skipped:
                        messages.warning(request, f"{len(skipped)} ligne(s) ignorée(s) (comptes déjà existants).")
                    if errors:
                        messages.error(request, f"{len(errors)} ligne(s) en erreur.")
                    if not created and not skipped and not errors:
                        messages.info(request, "Aucune ligne de données trouvée dans le fichier.")
            except StopIteration:
                pass
            except Exception as e:
                messages.error(request, f"Impossible de lire le fichier Excel: {str(e)}")

    context = {
        'admin': admin,
        'modules': modules,
        'created': created,
        'skipped': skipped,
        'errors': errors,
        'import_done': import_done,
    }

    return render(request, 'administrator/import_students.html', context)


@admin_required
def download_student_import_template(request):
    """Downloadable Excel template for the student import"""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Étudiants"
    sheet.append(['Prénom', 'Nom', 'Email', 'ID Étudiant', 'Année', 'Département'])
    sheet.append(['Amina', 'Benali', 'amina.benali@example.com', 'E12345', 3, 'Informatique'])
    sheet.append(['Youssef', 'El Amrani', 'youssef.elamrani@example.com', 'E12346', 4, 'Génie Civil'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_etudiants.xlsx"'
    workbook.save(response)
    return response

# Bulk teacher import from Excel (ROADMAP #12)

TEACHER_IMPORT_HEADER_ALIASES = {
    'prenom': 'first_name', 'prénom': 'first_name', 'first_name': 'first_name',
    'nom': 'last_name', 'last_name': 'last_name',
    'email': 'email', 'e-mail': 'email',
    'id enseignant': 'teacher_id', 'teacher_id': 'teacher_id', 'id': 'teacher_id',
    'departement': 'department', 'département': 'department', 'department': 'department',
}

TEACHER_IMPORT_REQUIRED_FIELDS = ['first_name', 'last_name', 'email', 'department']


@admin_required
def import_teachers(request):
    """Bulk-create teacher accounts from an Excel (.xlsx) roster"""
    from django.db import transaction
    from django.utils.crypto import get_random_string
    import openpyxl

    admin = get_object_or_404(AdminProfile, user=request.user)
    modules = Module.objects.filter(is_active=True).order_by('code')

    created = []
    skipped = []
    errors = []
    import_done = False

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        module_id = request.POST.get('module_id', '')

        module = None
        if module_id:
            module = get_object_or_404(Module, id=module_id)

        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel (.xlsx).")
        elif not excel_file.name.lower().endswith('.xlsx'):
            messages.error(request, "Format non supporté. Utilisez un fichier .xlsx (modèle téléchargeable ci-dessous).")
        else:
            try:
                workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet = workbook.active

                rows = sheet.iter_rows(values_only=True)
                header_row = next(rows, None)
                if header_row is None:
                    messages.error(request, "Le fichier est vide.")
                    raise StopIteration

                column_map = {}
                for index, header in enumerate(header_row):
                    if header is None:
                        continue
                    key = str(header).strip().lower()
                    if key in TEACHER_IMPORT_HEADER_ALIASES:
                        column_map[index] = TEACHER_IMPORT_HEADER_ALIASES[key]

                missing_columns = [f for f in TEACHER_IMPORT_REQUIRED_FIELDS if f not in column_map.values()]
                if missing_columns:
                    messages.error(
                        request,
                        "Colonnes manquantes dans le fichier: Prénom, Nom, Email, Département sont requises "
                        "(ID Enseignant optionnel). Téléchargez le modèle pour voir le format attendu."
                    )
                else:
                    for row_number, row in enumerate(rows, start=2):
                        if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
                            continue

                        data = {}
                        for index, field in column_map.items():
                            value = row[index] if index < len(row) else None
                            data[field] = str(value).strip() if value is not None else ''

                        missing = [f for f in TEACHER_IMPORT_REQUIRED_FIELDS if not data.get(f)]
                        if missing:
                            errors.append({'row': row_number, 'reason': "Valeurs manquantes (Prénom, Nom, Email et Département sont obligatoires)"})
                            continue

                        email = data['email'].lower()
                        teacher_id = data.get('teacher_id', '')

                        if User.objects.filter(email__iexact=email).exists():
                            skipped.append({'row': row_number, 'identity': email, 'reason': "Email déjà utilisé"})
                            continue
                        if teacher_id and TeacherProfile.objects.filter(teacher_id=teacher_id).exists():
                            skipped.append({'row': row_number, 'identity': teacher_id, 'reason': "ID enseignant déjà utilisé"})
                            continue

                        password = get_random_string(10)
                        try:
                            with transaction.atomic():
                                user = User.objects.create_user(
                                    username=_generate_unique_username(email),
                                    email=email,
                                    password=password,
                                    first_name=data['first_name'],
                                    last_name=data['last_name'],
                                    is_active=True,
                                )
                                teacher = TeacherProfile.objects.create(
                                    user=user,
                                    teacher_id=teacher_id or None,
                                    department=data['department'],
                                )
                                if module:
                                    ModuleAssignment.objects.create(
                                        teacher=teacher, module=module,
                                        assigned_by=request.user, is_active=True,
                                    )
                        except Exception as e:
                            errors.append({'row': row_number, 'reason': f"Erreur lors de la création: {str(e)}"})
                            continue

                        created.append({
                            'name': f"{data['first_name']} {data['last_name']}",
                            'teacher_id': teacher_id or '—',
                            'email': email,
                            'username': user.username,
                            'password': password,
                        })

                    import_done = True
                    if created:
                        assigned_note = f" et assignés au module {module.code}" if module else ""
                        messages.success(
                            request,
                            f"{len(created)} enseignant(s) créé(s){assigned_note}. "
                            "Notez les mots de passe ci-dessous: ils ne seront plus affichés."
                        )
                    if skipped:
                        messages.warning(request, f"{len(skipped)} ligne(s) ignorée(s) (comptes déjà existants).")
                    if errors:
                        messages.error(request, f"{len(errors)} ligne(s) en erreur.")
                    if not created and not skipped and not errors:
                        messages.info(request, "Aucune ligne de données trouvée dans le fichier.")
            except StopIteration:
                pass
            except Exception as e:
                messages.error(request, f"Impossible de lire le fichier Excel: {str(e)}")

    context = {
        'admin': admin,
        'modules': modules,
        'created': created,
        'skipped': skipped,
        'errors': errors,
        'import_done': import_done,
    }

    return render(request, 'administrator/import_teachers.html', context)


@admin_required
def download_teacher_import_template(request):
    """Downloadable Excel template for the teacher import"""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Enseignants"
    sheet.append(['Prénom', 'Nom', 'Email', 'ID Enseignant', 'Département'])
    sheet.append(['Karim', 'Alaoui', 'karim.alaoui@example.com', 'T1001', 'Informatique'])
    sheet.append(['Salma', 'Bennis', 'salma.bennis@example.com', '', 'Génie Civil'])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_enseignants.xlsx"'
    workbook.save(response)
    return response


# Account activation toggle (ROADMAP #13)

@admin_required
def toggle_user_active(request, user_id):
    """Activate/deactivate a user account (students and teachers only)."""
    if request.method != 'POST':
        return redirect('administrator:users_list')

    target = get_object_or_404(User, id=user_id)

    # Guards: never touch yourself or another administrator
    if target == request.user:
        messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
        return redirect('administrator:users_list')
    if AdminProfile.objects.filter(user=target).exists():
        messages.error(request, "Les comptes administrateur ne peuvent pas être désactivés ici.")
        return redirect('administrator:users_list')

    target.is_active = not target.is_active
    target.save(update_fields=['is_active'])

    name = target.get_full_name() or target.username
    if target.is_active:
        messages.success(request, f"Le compte de {name} a été réactivé.")
    else:
        messages.success(request, f"Le compte de {name} a été désactivé. Il ne peut plus se connecter.")

    # Back to the right tab
    user_type = 'teachers' if TeacherProfile.objects.filter(user=target).exists() else 'students'
    return redirect(f"/administrator/users/list/?type={user_type}")


# Optional Monitoring/Admin Views

@admin_required
def recent_module_activity(request):
    """Monitor recent module creation activity"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Get recently created modules (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_modules = Module.objects.filter(
        created_at__gte=thirty_days_ago
    ).select_related('primary_teacher__user').order_by('-created_at')
    
    # Separate teacher-created vs admin-created
    teacher_created = recent_modules.filter(created_by_teacher=True)
    admin_created = recent_modules.filter(created_by_teacher=False)
    
    # Get most active teachers
    active_teachers = TeacherProfile.objects.filter(
        primary_modules__created_at__gte=thirty_days_ago,
        primary_modules__created_by_teacher=True
    ).annotate(
        modules_created=Count('primary_modules')
    ).order_by('-modules_created')[:5]
    
    context = {
        'admin': admin,
        'recent_modules': recent_modules,
        'teacher_created': teacher_created,
        'admin_created': admin_created,
        'active_teachers': active_teachers,
        'total_recent': recent_modules.count(),
    }
    
    return render(request, 'administrator/recent_module_activity.html', context)


@admin_required  
def module_oversight(request):
    """Admin can review and manage all modules (without approval bottleneck)"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    
    # Get all modules with filters
    modules = Module.objects.select_related('primary_teacher__user').order_by('-created_at')
    
    # Apply filters
    creation_source = request.GET.get('source', '')
    teacher_filter = request.GET.get('teacher', '')
    year_filter = request.GET.get('year', '')
    
    if creation_source == 'teacher':
        modules = modules.filter(created_by_teacher=True)
    elif creation_source == 'admin':
        modules = modules.filter(created_by_teacher=False)
        
    if teacher_filter:
        modules = modules.filter(primary_teacher_id=teacher_filter)
        
    if year_filter:
        modules = modules.filter(academic_year=year_filter)
    
    # Calculate stats
    total_modules = modules.count()
    teacher_created_count = modules.filter(created_by_teacher=True).count()
    admin_created_count = modules.filter(created_by_teacher=False).count()
    
    # Pagination
    paginator = Paginator(modules, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # For filters
    teachers = TeacherProfile.objects.filter(
        primary_modules__isnull=False
    ).distinct().select_related('user').order_by('user__first_name')
    
    years = Module.objects.values_list('academic_year', flat=True).distinct()
    
    context = {
        'admin': admin,
        'page_obj': page_obj,
        'teachers': teachers,
        'years': years,
        'creation_source': creation_source,
        'teacher_filter': teacher_filter,
        'year_filter': year_filter,
        'total_modules': total_modules,
        'teacher_created_count': teacher_created_count,
        'admin_created_count': admin_created_count,
        # Admin can still deactivate problematic modules if needed
        'can_manage': True,
    }
    
    return render(request, 'administrator/module_oversight.html', context)


# Optional: Allow admin to deactivate modules if needed (rare cases)
@admin_required
def toggle_module_status(request, module_id):
    """Admin can activate/deactivate modules if needed"""
    admin = get_object_or_404(AdminProfile, user=request.user)
    module = get_object_or_404(Module, id=module_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        reason = request.POST.get('reason', '').strip()
        
        if action == 'deactivate' and module.is_active:
            module.is_active = False
            module.save()
            
            # Log the action
            messages.success(
                request, 
                f"Module '{module.code}' désactivé. Raison: {reason or 'Non spécifiée'}"
            )
            
        elif action == 'activate' and not module.is_active:
            module.is_active = True
            module.save()
            
            messages.success(request, f"Module '{module.code}' réactivé.")
        
        return redirect('administrator:module_oversight')
    
    context = {
        'admin': admin,
        'module': module,
    }
    
    return render(request, 'administrator/module_status_confirm.html', context)