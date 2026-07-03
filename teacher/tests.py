import shutil
import tempfile
from datetime import date

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse

from .models import Module, ModuleAssignment, TeacherProfile, ProjectEvaluation, EvaluationCriterion
from student.models import (
    DeliverableComment,
    Notification,
    Project,
    ProjectComment,
    ProjectDeliverable,
    StudentProfile,
)

TEMP_MEDIA_ROOT = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=TEMP_MEDIA_ROOT)
class DeliverableFeedbackTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEMP_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        teacher_user = User.objects.create_user(username='prof', password='profpass')
        self.teacher = TeacherProfile.objects.create(user=teacher_user, teacher_id='T001')

        student_user = User.objects.create_user(username='etudiant', password='etupass')
        self.student = StudentProfile.objects.create(
            user=student_user, student_id='E001', year_of_study=3, department='Informatique'
        )

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')
        ModuleAssignment.objects.create(teacher=self.teacher, module=self.module, is_active=True)

        self.project = Project.objects.create(
            title='Projet Test',
            description='Description',
            project_type='module',
            student=self.student,
            module=self.module,
            start_date=date(2026, 1, 1),
            status='submitted',
        )

        self.deliverable = ProjectDeliverable.objects.create(
            project=self.project,
            file=SimpleUploadedFile('rapport.pdf', b'contenu'),
            file_type='report',
            name='Rapport final',
        )

    def login_teacher(self):
        self.client.login(username='prof', password='profpass')

    def login_student(self):
        self.client.login(username='etudiant', password='etupass')

    def test_teacher_requests_revision(self):
        self.login_teacher()
        response = self.client.post(
            reverse('teacher:request_revision', args=[self.project.id]),
            {'content': 'Veuillez compléter la section 3 du rapport.'},
        )
        self.assertEqual(response.status_code, 302)

        self.project.refresh_from_db()
        self.assertEqual(self.project.status, 'revision_requested')
        self.assertTrue(
            ProjectComment.objects.filter(
                project=self.project, content__contains='Veuillez compléter la section 3'
            ).exists()
        )
        self.assertTrue(
            Notification.objects.filter(recipient=self.student, project=self.project).exists()
        )

    def test_request_revision_requires_content(self):
        self.login_teacher()
        self.client.post(reverse('teacher:request_revision', args=[self.project.id]), {'content': '  '})

        self.project.refresh_from_db()
        self.assertEqual(self.project.status, 'submitted')

    def test_teacher_comments_on_deliverable(self):
        self.login_teacher()
        response = self.client.post(
            reverse('teacher:add_deliverable_comment', args=[self.deliverable.id]),
            {'content': 'Le rapport manque de sources.'},
        )
        self.assertEqual(response.status_code, 302)

        comment = DeliverableComment.objects.get(deliverable=self.deliverable)
        self.assertEqual(comment.author, self.teacher.user)
        self.assertTrue(comment.is_teacher_comment)
        self.assertTrue(
            Notification.objects.filter(
                recipient=self.student, notification_type='deliverable'
            ).exists()
        )

    def test_student_replies_on_deliverable(self):
        self.login_student()
        self.client.post(
            reverse('student:add_deliverable_comment', args=[self.deliverable.id]),
            {'content': 'Sources ajoutées dans la nouvelle version.'},
        )

        comment = DeliverableComment.objects.get(deliverable=self.deliverable)
        self.assertEqual(comment.author, self.student.user)
        self.assertFalse(comment.is_teacher_comment)

    def test_other_student_cannot_comment_on_deliverable(self):
        outsider = User.objects.create_user(username='intrus', password='intruspass')
        StudentProfile.objects.create(
            user=outsider, student_id='E002', year_of_study=3, department='Informatique'
        )
        self.client.login(username='intrus', password='intruspass')

        self.client.post(
            reverse('student:add_deliverable_comment', args=[self.deliverable.id]),
            {'content': 'Je ne devrais pas pouvoir commenter.'},
        )
        self.assertEqual(DeliverableComment.objects.count(), 0)

    def test_student_can_resubmit_after_revision_request(self):
        self.project.status = 'revision_requested'
        self.project.save()

        self.login_student()
        response = self.client.post(reverse('student:project_submit', args=[self.project.id]))
        self.assertEqual(response.status_code, 302)

        self.project.refresh_from_db()
        self.assertEqual(self.project.status, 'submitted')


class GradingTests(TestCase):
    def setUp(self):
        teacher_user = User.objects.create_user(username='prof', password='profpass')
        self.teacher = TeacherProfile.objects.create(user=teacher_user, teacher_id='T001')

        student_user = User.objects.create_user(username='etudiant', password='etupass')
        self.student = StudentProfile.objects.create(
            user=student_user, student_id='E001', year_of_study=3, department='Informatique'
        )

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')
        ModuleAssignment.objects.create(teacher=self.teacher, module=self.module, is_active=True)

        self.project = Project.objects.create(
            title='Projet Test',
            description='Description',
            project_type='module',
            student=self.student,
            module=self.module,
            start_date=date(2026, 1, 1),
            status='submitted',
        )

    def login_teacher(self):
        self.client.login(username='prof', password='profpass')

    def grade_url(self):
        return reverse('teacher:grade_project', args=[self.project.id])

    def test_grade_page_renders(self):
        self.login_teacher()
        response = self.client.get(self.grade_url())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Noter le projet')

    def test_grading_with_validation(self):
        self.login_teacher()
        response = self.client.post(self.grade_url(), {
            'score': '15.5',
            'comments': 'Bon travail, présentation solide.',
            'validate_project': 'on',
            'criterion_name': ['Qualité technique', 'Respect des délais', ''],
            'criterion_score': ['16', '15', ''],
            'criterion_comment': ['Code propre', '', ''],
        })
        self.assertEqual(response.status_code, 302)

        evaluation = ProjectEvaluation.objects.get(project=self.project)
        self.assertEqual(float(evaluation.score), 15.5)
        self.assertEqual(evaluation.get_mention(), 'Bien')
        self.assertEqual(evaluation.criteria.count(), 2)  # empty row skipped

        self.project.refresh_from_db()
        self.assertEqual(self.project.status, 'validated')

        # Team notified with the grade
        self.assertTrue(
            Notification.objects.filter(
                recipient=self.student, message__contains='15.5'
            ).exists()
        )

    def test_grading_without_validation(self):
        self.login_teacher()
        self.client.post(self.grade_url(), {
            'score': '12',
            'comments': '',
        })
        self.project.refresh_from_db()
        self.assertEqual(self.project.status, 'submitted')
        self.assertTrue(ProjectEvaluation.objects.filter(project=self.project).exists())

    def test_invalid_score_rejected(self):
        self.login_teacher()
        self.client.post(self.grade_url(), {'score': '25'})
        self.assertFalse(ProjectEvaluation.objects.filter(project=self.project).exists())

        self.client.post(self.grade_url(), {'score': ''})
        self.assertFalse(ProjectEvaluation.objects.filter(project=self.project).exists())

    def test_regrade_replaces_rubric(self):
        self.login_teacher()
        self.client.post(self.grade_url(), {
            'score': '10',
            'criterion_name': ['Ancien critère'],
            'criterion_score': ['10'],
            'criterion_comment': [''],
        })
        self.client.post(self.grade_url(), {
            'score': '14,5',  # French comma decimal
            'criterion_name': ['Nouveau critère'],
            'criterion_score': ['14'],
            'criterion_comment': [''],
        })

        evaluation = ProjectEvaluation.objects.get(project=self.project)
        self.assertEqual(float(evaluation.score), 14.5)
        names = list(evaluation.criteria.values_list('name', flat=True))
        self.assertEqual(names, ['Nouveau critère'])

    def test_other_teacher_cannot_grade(self):
        other_user = User.objects.create_user(username='autre', password='autrepass')
        TeacherProfile.objects.create(user=other_user, teacher_id='T002')
        self.client.login(username='autre', password='autrepass')

        response = self.client.post(self.grade_url(), {'score': '18'})
        self.assertEqual(response.status_code, 404)
        self.assertFalse(ProjectEvaluation.objects.exists())

    def test_student_sees_grade_on_project_page(self):
        ProjectEvaluation.objects.create(project=self.project, teacher=self.teacher, score=16)
        self.client.login(username='etudiant', password='etupass')

        response = self.client.get(reverse('student:project_detail', args=[self.project.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '16')
        self.assertContains(response, 'Très bien')


class BulkActionTests(TestCase):
    def setUp(self):
        from django.utils import timezone
        from .models import ProjectAssignment

        teacher_user = User.objects.create_user(username='prof', password='profpass')
        self.teacher = TeacherProfile.objects.create(user=teacher_user, teacher_id='T001')

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')
        ModuleAssignment.objects.create(teacher=self.teacher, module=self.module, is_active=True)

        self.assignment = ProjectAssignment.objects.create(
            teacher=self.teacher,
            module=self.module,
            title='Devoir de test',
            description='Description',
            deadline=timezone.now() + timezone.timedelta(days=14),
            assignment_type='direct',
            status='published',
        )

        self.projects = []
        for i in range(3):
            student_user = User.objects.create_user(username=f'etu{i}', password='pass')
            student = StudentProfile.objects.create(
                user=student_user, student_id=f'E00{i}', year_of_study=3, department='Info'
            )
            self.projects.append(Project.objects.create(
                title=f'Projet {i}',
                description='d',
                project_type='module',
                student=student,
                module=self.module,
                start_date=date(2026, 1, 1),
                status='submitted',
                assignment_source='teacher_assigned',
                project_assignment=self.assignment,
            ))

    def bulk_url(self):
        return reverse('teacher:bulk_project_action', args=[self.assignment.id])

    def test_bulk_validate(self):
        self.client.login(username='prof', password='profpass')
        response = self.client.post(self.bulk_url(), {
            'action': 'validate',
            'project_ids': [p.id for p in self.projects[:2]],
        })
        self.assertEqual(response.status_code, 302)

        for project in self.projects[:2]:
            project.refresh_from_db()
            self.assertEqual(project.status, 'validated')
        self.projects[2].refresh_from_db()
        self.assertEqual(self.projects[2].status, 'submitted')

        # Each validated team was notified
        self.assertEqual(
            Notification.objects.filter(message__contains='validé').count(), 2
        )

    def test_bulk_revision_requires_content(self):
        self.client.login(username='prof', password='profpass')
        self.client.post(self.bulk_url(), {
            'action': 'revision',
            'project_ids': [self.projects[0].id],
            'content': '  ',
        })
        self.projects[0].refresh_from_db()
        self.assertEqual(self.projects[0].status, 'submitted')

    def test_bulk_revision(self):
        self.client.login(username='prof', password='profpass')
        self.client.post(self.bulk_url(), {
            'action': 'revision',
            'project_ids': [p.id for p in self.projects],
            'content': 'Complétez le rapport.',
        })
        for project in self.projects:
            project.refresh_from_db()
            self.assertEqual(project.status, 'revision_requested')
            self.assertTrue(
                ProjectComment.objects.filter(
                    project=project, content__contains='Complétez le rapport.'
                ).exists()
            )

    def test_bulk_skips_foreign_projects(self):
        """Projects that don't belong to the assignment are never touched."""
        outsider_user = User.objects.create_user(username='autre', password='pass')
        outsider = StudentProfile.objects.create(
            user=outsider_user, student_id='E999', year_of_study=3, department='Info'
        )
        foreign = Project.objects.create(
            title='Hors devoir', description='d', project_type='module',
            student=outsider, module=self.module,
            start_date=date(2026, 1, 1), status='submitted',
        )

        self.client.login(username='prof', password='profpass')
        self.client.post(self.bulk_url(), {
            'action': 'validate',
            'project_ids': [foreign.id],
        })
        foreign.refresh_from_db()
        self.assertEqual(foreign.status, 'submitted')

    def test_bulk_requires_own_assignment(self):
        other_user = User.objects.create_user(username='prof2', password='pass')
        TeacherProfile.objects.create(user=other_user, teacher_id='T002')
        self.client.login(username='prof2', password='pass')

        response = self.client.post(self.bulk_url(), {
            'action': 'validate',
            'project_ids': [self.projects[0].id],
        })
        self.assertEqual(response.status_code, 404)
        self.projects[0].refresh_from_db()
        self.assertEqual(self.projects[0].status, 'submitted')


class AnnouncementTests(TestCase):
    def setUp(self):
        from .models import ModuleEnrollment

        teacher_user = User.objects.create_user(username='prof', password='profpass')
        self.teacher = TeacherProfile.objects.create(user=teacher_user, teacher_id='T001')

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')
        ModuleAssignment.objects.create(teacher=self.teacher, module=self.module, is_active=True)

        # Two enrolled students, one inactive enrollment, one outsider
        self.enrolled = []
        for i in range(2):
            user = User.objects.create_user(username=f'etu{i}', password='pass', email=f'etu{i}@x.com')
            profile = StudentProfile.objects.create(
                user=user, student_id=f'E00{i}', year_of_study=3, department='Info'
            )
            ModuleEnrollment.objects.create(student=profile, module=self.module, is_active=True)
            self.enrolled.append(profile)

        inactive_user = User.objects.create_user(username='parti', password='pass')
        self.inactive = StudentProfile.objects.create(
            user=inactive_user, student_id='E098', year_of_study=3, department='Info'
        )
        ModuleEnrollment.objects.create(student=self.inactive, module=self.module, is_active=False)

        self.url = reverse('teacher:post_module_announcement', args=[self.module.id])

    def test_announcement_notifies_active_students_only(self):
        from .models import ModuleAnnouncement

        self.client.login(username='prof', password='profpass')
        response = self.client.post(self.url, {
            'title': 'Changement de salle',
            'content': 'Le cours de lundi aura lieu en B204.',
        })
        self.assertEqual(response.status_code, 302)

        announcement = ModuleAnnouncement.objects.get(module=self.module)
        self.assertEqual(announcement.title, 'Changement de salle')
        self.assertEqual(announcement.teacher, self.teacher)

        notifications = Notification.objects.filter(notification_type='announcement')
        self.assertEqual(notifications.count(), 2)
        recipients = {n.recipient_id for n in notifications}
        self.assertEqual(recipients, {s.id for s in self.enrolled})
        self.assertNotIn(self.inactive.id, recipients)
        self.assertIn('B204', notifications.first().message)

    def test_announcement_requires_title_and_content(self):
        from .models import ModuleAnnouncement

        self.client.login(username='prof', password='profpass')
        self.client.post(self.url, {'title': '  ', 'content': ''})
        self.assertEqual(ModuleAnnouncement.objects.count(), 0)
        self.assertEqual(Notification.objects.count(), 0)

    def test_announcement_requires_module_access(self):
        from .models import ModuleAnnouncement

        other = User.objects.create_user(username='prof2', password='pass')
        TeacherProfile.objects.create(user=other, teacher_id='T002')
        self.client.login(username='prof2', password='pass')

        self.client.post(self.url, {'title': 'Intrusion', 'content': 'x'})
        self.assertEqual(ModuleAnnouncement.objects.count(), 0)

    def test_announcements_listed_on_module_page(self):
        from .models import ModuleAnnouncement

        ModuleAnnouncement.objects.create(
            module=self.module, teacher=self.teacher,
            title='Rappel examen', content='Réviser les chapitres 1 à 4.',
        )
        self.client.login(username='prof', password='profpass')
        response = self.client.get(reverse('teacher:module_detail', args=[self.module.id]))
        self.assertContains(response, 'Rappel examen')


class TeacherExportTests(TestCase):
    def setUp(self):
        from django.utils import timezone
        from .models import ProjectAssignment, ProjectEvaluation, ModuleEnrollment

        teacher_user = User.objects.create_user(username='prof', password='profpass')
        self.teacher = TeacherProfile.objects.create(user=teacher_user, teacher_id='T001')

        self.module = Module.objects.create(name='Programmation Web', code='WEB301')
        ModuleAssignment.objects.create(teacher=self.teacher, module=self.module, is_active=True)

        student_user = User.objects.create_user(
            username='etu', password='pass', first_name='Amina', last_name='Benali'
        )
        self.student = StudentProfile.objects.create(
            user=student_user, student_id='E001', year_of_study=3, department='Info'
        )
        ModuleEnrollment.objects.create(student=self.student, module=self.module, is_active=True)

        self.assignment = ProjectAssignment.objects.create(
            teacher=self.teacher, module=self.module,
            title='Devoir export', description='d',
            deadline=timezone.now() + timezone.timedelta(days=14),
            assignment_type='direct', status='published',
        )
        self.project = Project.objects.create(
            title='Projet noté', description='d', project_type='module',
            student=self.student, module=self.module,
            start_date=date(2026, 1, 1), status='validated',
            assignment_source='teacher_assigned', project_assignment=self.assignment,
        )
        ProjectEvaluation.objects.create(project=self.project, teacher=self.teacher, score=15)

        self.client.login(username='prof', password='profpass')

    def read_workbook(self, response):
        from io import BytesIO
        from openpyxl import load_workbook
        return load_workbook(BytesIO(response.content))

    def test_module_roster_export(self):
        response = self.client.get(reverse('teacher:export_module_roster', args=[self.module.id]))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

        sheet = self.read_workbook(response).active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertIn('Benali', [row[1] for row in rows[1:]])

    def test_assignment_results_export_includes_grade(self):
        response = self.client.get(
            reverse('teacher:export_assignment_results', args=[self.assignment.id])
        )
        self.assertEqual(response.status_code, 200)

        sheet = self.read_workbook(response).active
        rows = list(sheet.iter_rows(values_only=True))
        data = rows[1]
        self.assertEqual(data[0], 'Projet noté')
        self.assertEqual(data[8], 15.0)   # Note /20
        self.assertEqual(data[9], 'Bien')  # Mention

    def test_roster_export_requires_module_access(self):
        other_user = User.objects.create_user(username='prof2', password='pass')
        TeacherProfile.objects.create(user=other_user, teacher_id='T002')
        self.client.login(username='prof2', password='pass')

        response = self.client.get(reverse('teacher:export_module_roster', args=[self.module.id]))
        self.assertEqual(response.status_code, 302)  # bounced with an error message

    def test_results_export_requires_own_assignment(self):
        other_user = User.objects.create_user(username='prof2', password='pass')
        TeacherProfile.objects.create(user=other_user, teacher_id='T002')
        self.client.login(username='prof2', password='pass')

        response = self.client.get(
            reverse('teacher:export_assignment_results', args=[self.assignment.id])
        )
        self.assertEqual(response.status_code, 404)
